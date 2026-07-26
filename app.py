from flask import Flask, render_template, request, redirect, url_for, jsonify
import threading
import config
from config import APP_HOST, APP_PORT, APP_DEBUG
from db import init_db
from modules.connections import (
    list_connections,
    create_connection,
    delete_connection,
    test_gp_connection,
)

from modules.gpcopy_sync import (
    preview_gpcopy_sync,
    run_gpcopy_sync_job,
)

import io
import os
import sqlite3
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from modules.vacuum_analyze import run_vacuum_analyze_job
from job_manager import (
    create_job,
    create_job_items,
    get_active_jobs,
    get_job,
    get_job_items,
    get_latest_job,
    list_recent_jobs,
    mark_interrupted_jobs_on_startup,
    request_stop_job,
    set_stop_flag,
)
from modules.skew_analyzer import (
    analyze_tables_skew,
    get_last_skew_results,
    run_skew_job,
    get_skew_results_by_job,
    get_skew_summary_by_job,
    get_skew_result_segments,
    get_latest_problem_skew_results,
)
from modules.object_tree import get_object_tree

from modules.reorganize import (
    get_reorganize_targets,
    run_reorganize_job,
    get_distribution_recommendation,
    apply_distribution_and_reorganize,
)


from modules.gpcopy import (
    run_gpcopy_job,
    build_gpcopy_date_include_json_preview,
    build_retry_config,
    get_date_columns_for_table,
    get_gpcopy_date_columns,
)

from modules.dashboard import get_session_limits_stats

import json as _json
from datetime import datetime as _datetime

import scheduler as gpm_scheduler
import scheduler_store
from modules.date_window import resolve_date_window

# Реестр раннеров для планировщика (spec §5): job_type -> существующая функция.
gpm_scheduler.register_runner("gpcopy", run_gpcopy_job)
gpm_scheduler.register_runner("gpcopy_date", run_gpcopy_job)
gpm_scheduler.register_runner("vacuum", run_vacuum_analyze_job)
gpm_scheduler.register_runner("skew", run_skew_job)
gpm_scheduler.register_runner("reorganize", run_reorganize_job)

from modules.gpcopy_increment import (
    run_gpcopy_increment_job,
    get_dest_watermark,
    build_increment_items,
)
from modules.gpcopy_partition import (
    run_gpcopy_partition_diff_job,
    diff_partitions,
)

gpm_scheduler.register_runner("gpcopy_increment", run_gpcopy_increment_job)
gpm_scheduler.register_runner("gpcopy_partition_diff", run_gpcopy_partition_diff_job)
gpm_scheduler.register_runner("gpcopy_sync", run_gpcopy_sync_job)

from modules.sync_transport import pick_transport, run_copy_pipe_job

gpm_scheduler.register_runner("copy_pipe", run_copy_pipe_job)

from modules.gpbackup import (
    list_backups as list_backup_records,
    run_gpbackup_job,
    run_gprestore_job,
)

gpm_scheduler.register_runner("gpbackup", run_gpbackup_job)
gpm_scheduler.register_runner("gprestore", run_gprestore_job)


app = Flask(__name__)


@app.route("/")
@app.route("/dashboard")
def dashboard_page():
    connections = list_connections()
    last_results = get_last_skew_results(limit=1000)
    latest_skew_job = get_latest_job("skew")

    skew_summary = build_skew_dashboard_summary(last_results)

    return render_template(
        "dashboard.html",
        connections=connections,
        last_results=last_results,
        latest_skew_job=latest_skew_job,
        skew_summary=skew_summary,
    )


@app.route("/connections")
def connections_page():
    connections = list_connections()
    return render_template(
        "connections.html",
        connections=connections,
    )


@app.route("/connections/add", methods=["POST"])
def add_connection():
    try:
        create_connection(request.form.to_dict())
        return redirect(url_for("connections_page"))
    except Exception as e:
        connections = list_connections()
        return render_template(
            "connections.html",
            connections=connections,
            error=str(e),
        )


@app.route("/connections/delete/<int:connection_id>", methods=["POST"])
def remove_connection(connection_id):
    delete_connection(connection_id)
    return redirect(url_for("connections_page"))


@app.route("/objects")
def objects_page():
    connections = list_connections()
    return render_template(
        "objects.html",
        connections=connections,
    )


@app.route("/api/connections")
def api_connections():
    return jsonify(
        {
            "ok": True,
            "connections": list_connections(),
        }
    )


@app.route("/api/connections/<int:connection_id>/test", methods=["POST"])
def api_test_connection(connection_id):
    result = test_gp_connection(connection_id)
    return jsonify(result)


@app.route("/api/objects/tree")
def api_objects_tree():
    connection_id = request.args.get("connection_id", type=int)

    if not connection_id:
        return jsonify(
            {
                "ok": False,
                "message": "connection_id обязателен",
            }
        ), 400

    try:
        tree = get_object_tree(connection_id)
        return jsonify(
            {
                "ok": True,
                "tree": tree,
            }
        )
    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "message": str(e),
            }
        ), 500


@app.route("/skew")
def skew_page():
    return redirect(url_for("maintenance_page"))


@app.route("/api/skew/analyze", methods=["POST"])
def api_skew_analyze():
    data = request.get_json(silent=True) or {}

    connection_id = data.get("connection_id")
    tables = data.get("tables") or []

    if not connection_id:
        return jsonify(
            {
                "ok": False,
                "message": "connection_id обязателен",
            }
        ), 400

    if not tables:
        return jsonify(
            {
                "ok": False,
                "message": "Не выбраны таблицы",
            }
        ), 400

    try:
        results = analyze_tables_skew(
            connection_id=int(connection_id),
            tables=tables,
        )

        return jsonify(
            {
                "ok": True,
                "results": results,
            }
        )

    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "message": str(e),
            }
        ), 500


@app.route("/api/skew/results")
def api_skew_results():
    limit = request.args.get("limit", 100, type=int)

    return jsonify(
        {
            "ok": True,
            "results": get_last_skew_results(limit),
        }
    )

@app.route("/api/skew/start", methods=["POST"])
def api_skew_start():
    data = request.get_json(silent=True) or {}

    connection_id = (
        data.get("connection_id")
        or data.get("source_connection_id")
        or data.get("connectionId")
    )

    tables = data.get("tables") or data.get("selected_tables") or []

    if not connection_id:
        return jsonify({
            "ok": False,
            "message": "connection_id is required",
        }), 400

    if not tables:
        return jsonify({
            "ok": False,
            "message": "tables is empty",
        }), 400

    try:
        job_id = create_job(
            job_type="skew",
            connection_id=int(connection_id),
            config={
                "connection_id": int(connection_id),
                "tables": tables,
            },
        )

        create_job_items(
            job_id=job_id,
            items=[
                {
                    "schema_name": item.get("schema") or item.get("schema_name"),
                    "table_name": item.get("table") or item.get("table_name"),
                    "action": "SKEW",
                }
                for item in tables
                if (item.get("schema") or item.get("schema_name"))
                and (item.get("table") or item.get("table_name"))
            ],
        )

        import threading

        threading.Thread(
            target=run_skew_job,
            args=(job_id,),
            daemon=True,
        ).start()

        return jsonify({
            "ok": True,
            "job_id": job_id,
            "message": "Skew job started",
        })

    except Exception as e:
        import traceback

        return jsonify({
            "ok": False,
            "message": str(e),
            "traceback": traceback.format_exc(),
        }), 500

@app.route("/api/jobs/<int:job_id>")
def api_get_job(job_id):
    job = get_job(job_id)

    if not job:
        return jsonify(
            {
                "ok": False,
                "message": "Job not found",
            }
        ), 404

    return jsonify(
        {
            "ok": True,
            "job": job,
        }
    )

@app.route("/api/gpcopy/precheck", methods=["POST"])
def api_gpcopy_precheck():
    """Предпроверка DDL: сравнение колонок источника и приёмника."""
    data = request.get_json(silent=True) or {}
    tables = data.get("tables") or []

    if not data.get("source_connection_id") or not data.get("dest_connection_id"):
        return jsonify({"ok": False,
                        "message": "source и dest подключения обязательны"}), 400

    if not tables:
        return jsonify({"ok": False, "message": "Не выбраны таблицы"}), 400

    try:
        from modules.ddl_check import precheck_tables

        result = precheck_tables(
            data["source_connection_id"], data["dest_connection_id"], tables,
        )
        return jsonify({"ok": True, **result})
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:2000]}), 500


@app.route("/api/gpcopy/add-columns", methods=["POST"])
def api_gpcopy_add_columns():
    """Досоздать в приёмнике колонки, которых не хватает до источника."""
    data = request.get_json(silent=True) or {}
    tables = data.get("tables") or []

    if not data.get("dest_connection_id") or not tables:
        return jsonify({"ok": False,
                        "message": "dest_connection_id и tables обязательны"}), 400

    try:
        from modules.ddl_check import add_missing_columns

        results = add_missing_columns(data["dest_connection_id"], tables)
        return jsonify({
            "ok": True,
            "results": results,
            "added": sum(r["added"] for r in results),
            "failed": sum(1 for r in results if not r["ok"]),
        })
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:2000]}), 500


@app.route("/api/gpcopy/retry-failed", methods=["POST"])
def api_gpcopy_retry_failed():
    """Новая задача только по упавшим партициям исходной gpcopy-задачи."""
    data = request.get_json(silent=True) or {}
    job_id = data.get("job_id")

    if not job_id:
        return jsonify({"ok": False, "message": "job_id обязателен"}), 400

    job = get_job(int(job_id))

    if not job:
        return jsonify({"ok": False, "message": "Задача не найдена"}), 404

    if job.get("job_type") != "gpcopy":
        return jsonify({
            "ok": False,
            "message": "Дозагрузка упавших доступна только для gpcopy-задач",
        }), 400

    config = _json.loads(job.get("config_json") or "{}")
    failed_leaves = [tuple(p) for p in (config.get("failed_leaves") or [])]

    if not failed_leaves:
        return jsonify({
            "ok": False,
            "message": "У задачи нет сохранённого списка упавших партиций "
                       "(она запускалась до этой версии или упала целиком) — "
                       "перезапусти копирование обычным способом",
        }), 400

    try:
        retry_config = build_retry_config(config, failed_leaves)
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400

    retry_config["retry_of_job_id"] = int(job_id)

    new_job_id = create_job(
        job_type="gpcopy",
        connection_id=job.get("connection_id"),
        config=retry_config,
    )

    create_job_items(
        job_id=new_job_id,
        items=[
            {
                "schema_name": schema,
                "table_name": table,
                "action": "GPCOPY RETRY",
            }
            for schema, table in failed_leaves
        ],
    )

    threading.Thread(
        target=run_gpcopy_job,
        args=(new_job_id,),
        daemon=True,
    ).start()

    return jsonify({
        "ok": True,
        "job_id": new_job_id,
        "total_items": len(failed_leaves),
        "message": "Дозагрузка {} партиций запущена".format(len(failed_leaves)),
    })


@app.route("/api/jobs/<int:job_id>/status")
def api_job_status(job_id):
    job = get_job(job_id)

    if not job:
        return jsonify(
            {
                "ok": False,
                "message": "Job not found",
            }
        ), 404

    items = get_job_items(job_id)

    total = len(items)
    done = len([i for i in items if i.get("status") == "done"])
    failed = len([i for i in items if i.get("status") == "failed"])
    skipped = len([i for i in items if i.get("status") == "skipped"])
    running = len([i for i in items if i.get("status") == "running"])

    finished = done + failed + skipped

    percent = 0

    if total > 0:
        percent = round((finished * 100.0) / total, 2)

    return jsonify(
        {
            "ok": True,
            "job": job,
            "items": items,
            "summary": {
                "total": total,
                "done": done,
                "failed": failed,
                "skipped": skipped,
                "running": running,
                "finished": finished,
                "percent": percent,
            },
        }
    )

@app.route("/api/jobs/<int:job_id>/items")
def api_get_job_items(job_id):
    return jsonify(
        {
            "ok": True,
            "items": get_job_items(job_id),
        }
    )


@app.route("/api/jobs/<int:job_id>/stop", methods=["POST"])
def api_stop_job(job_id):
    try:
        job = get_job(job_id)

        if not job:
            return jsonify({
                "ok": False,
                "message": "Job not found",
            }), 404

        status = (
            job.get("status")
            if isinstance(job, dict)
            else get_item_value(job, "status")
        )

        if status in ("done", "failed", "cancelled", "interrupted"):
            return jsonify({
                "ok": True,
                "message": "Job already finished",
                "job_id": job_id,
                "status": status,
            })

        # Важно: ставим stop flag, а не убиваем Flask/thread напрямую
        request_stop_job(job_id)

        try:
            mark_job_stopping(job_id)
        except Exception:
            # Если такой функции нет — не падаем
            pass

        return jsonify({
            "ok": True,
            "job_id": job_id,
            "message": "Stop requested",
        })

    except Exception as e:
        import traceback

        return jsonify({
            "ok": False,
            "message": str(e),
            "traceback": traceback.format_exc(),
        }), 500


@app.route("/api/jobs/<int:job_id>/skew-results")
def api_get_job_skew_results(job_id):
    try:
        results = get_skew_results_by_job(job_id)
        summary = get_skew_summary_by_job(job_id)

        return jsonify(
            {
                "ok": True,
                "results": results,
                "summary": summary,
            }
        )
    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "message": str(e),
            }
        ), 500


@app.route("/api/jobs/latest/skew")
def api_get_latest_skew_job():
    job = get_latest_job("skew")

    if not job:
        return jsonify(
            {
                "ok": False,
                "message": "No skew job found",
            }
        ), 404

    return jsonify(
        {
            "ok": True,
            "job": job,
        }
    )


@app.route("/api/jobs/active")
def api_get_active_jobs():
    job_type = request.args.get("job_type")

    return jsonify(
        {
            "ok": True,
            "jobs": get_active_jobs(job_type),
        }
    )


@app.route("/api/jobs/recent")
def api_jobs_recent():
    types_arg = (request.args.get("types") or "").strip()
    job_types = [t.strip() for t in types_arg.split(",") if t.strip()] or None
    limit = request.args.get("limit", 20, type=int)

    return jsonify({
        "ok": True,
        "jobs": list_recent_jobs(job_types, limit),
    })


@app.route("/api/skew-results/<int:result_id>/segments")
def api_get_skew_result_segments(result_id):
    try:
        data = get_skew_result_segments(result_id)

        if not data:
            return jsonify(
                {
                    "ok": False,
                    "message": "Skew result not found",
                }
            ), 404

        return jsonify(
            {
                "ok": True,
                "result": data["result"],
                "segments": data["segments"],
            }
        )

    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "message": str(e),
            }
        ), 500


@app.route("/reorganize")
def reorganize_page():
    return redirect(url_for("maintenance_page"))


@app.route("/api/reorganize/start", methods=["POST"])
def api_start_reorganize():
    try:
        payload = request.get_json() or {}

        connection_id = payload.get("connection_id")
        selected_tables = payload.get("tables") or []

        if not connection_id:
            return jsonify(
                {
                    "ok": False,
                    "message": "connection_id is required",
                }
            ), 400

        if not selected_tables:
            return jsonify(
                {
                    "ok": False,
                    "message": "No tables selected",
                }
            ), 400

        expanded_tables = []

        for item in selected_tables:
            schema_name = item.get("schema")
            table_name = item.get("table")

            if not schema_name or not table_name:
                continue

            targets = get_reorganize_targets(
                connection_id=connection_id,
                schema_name=schema_name,
                table_name=table_name,
            )

            for target in targets:
                key = "{}.{}".format(
                    target["schema_name"],
                    target["table_name"],
                )

                expanded_tables.append(
                    {
                        "schema": target["schema_name"],
                        "table": target["table_name"],
                        "full_name": key,
                    }
                )

        unique_tables = []
        seen = set()

        for item in expanded_tables:
            key = "{}.{}".format(item["schema"], item["table"])

            if key in seen:
                continue

            seen.add(key)
            unique_tables.append(item)

        if not unique_tables:
            return jsonify(
                {
                    "ok": False,
                    "message": "No reorganize targets found",
                }
            ), 400  
      
        job_id = create_job(
            job_type="reorganize",
            connection_id=connection_id,
            config={
                "source": "web",
                "selected_tables": selected_tables,
                "expanded_tables": unique_tables,
            },
        )       

        create_job_items(
            job_id=job_id,
            items=[
                {
                    "schema_name": item["schema"],
                    "table_name": item["table"],
                    "action": "REORGANIZE",
                }
                for item in unique_tables
            ],
        )  

        #run_background_job(
        #    target=run_reorganize_job,
        #    args=(job_id,),
        #)

        threading.Thread(
            target=run_reorganize_job,
            args=(job_id,),
            daemon=True
        ).start()

        return jsonify(
            {
                "ok": True,
                "job_id": job_id,
                "total_items": len(unique_tables),
            }
        )

    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "message": str(e),
            }
        ), 500    


@app.route("/api/reorganize/recommendation", methods=["POST"])
def api_reorganize_recommendation():
    try:
        data = request.get_json(force=True)

        connection_id = data.get("connection_id")
        schema_name = data.get("schema_name")
        table_name = data.get("table_name")

        if not connection_id or not schema_name or not table_name:
            return jsonify(
                {
                    "ok": False,
                    "message": "connection_id, schema_name, table_name are required",
                }
            ), 400

        result = get_distribution_recommendation(
            connection_id=connection_id,
            schema_name=schema_name,
            table_name=table_name,
        )

        return jsonify(result)

    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "message": str(e),
            }
        ), 500


@app.route("/api/reorganize/apply-distribution", methods=["POST"])
def api_reorganize_apply_distribution():
    try:
        data = request.get_json(force=True)

        connection_id = data.get("connection_id")
        schema_name = data.get("schema_name")
        table_name = data.get("table_name")
        distribution_type = data.get("distribution_type")
        columns = data.get("columns") or []

        if not connection_id or not schema_name or not table_name or not distribution_type:
            return jsonify(
                {
                    "ok": False,
                    "message": "connection_id, schema_name, table_name, distribution_type are required",
                }
            ), 400

        result = apply_distribution_and_reorganize(
            connection_id=connection_id,
            schema_name=schema_name,
            table_name=table_name,
            distribution_type=distribution_type,
            columns=columns,
        )

        return jsonify(result)

    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "message": str(e),
            }
        ), 500

@app.route("/maintenance")
def maintenance_page():
    connections = list_connections()

    last_results = get_last_skew_results(limit=1000)
    problem_skew_results = get_latest_problem_skew_results(limit=500)

    latest_skew_job = get_latest_job("skew")
    latest_reorganize_job = get_latest_job("reorganize")

    return render_template(
        "maintenance.html",
        connections=connections,
        last_results=last_results,
        problem_skew_results=problem_skew_results,
        latest_skew_job=latest_skew_job,
        latest_reorganize_job=latest_reorganize_job,
    )

def build_skew_dashboard_summary(results):
    summary = {
        "total": 0,
        "ok": 0,
        "warning": 0,
        "critical": 0,
        "empty": 0,
        "failed": 0,
        "interrupted": 0,
        "max_skew": 0,
        "avg_skew": 0,
    }

    if not results:
        return summary

    total_skew = 0

    for r in results:
        summary["total"] += 1

        status = str(r.get("status") or "").upper()

        if status == "OK":
            summary["ok"] += 1
        elif status == "WARNING":
            summary["warning"] += 1
        elif status == "CRITICAL":
            summary["critical"] += 1
        elif status == "EMPTY":
            summary["empty"] += 1
        elif status == "FAILED":
            summary["failed"] += 1
        elif status == "INTERRUPTED":
            summary["interrupted"] += 1

        skew = float(r.get("skew_ratio") or 0)
        total_skew += skew

        if skew > summary["max_skew"]:
            summary["max_skew"] = skew

    summary["avg_skew"] = round(total_skew / summary["total"], 4)

    return summary

@app.route("/vacuum")
def vacuum_page():
    connections = list_connections()

    latest_vacuum_job = get_latest_job("vacuum_analyze")

    return render_template(
        "vacuum.html",
        connections=connections,
        latest_vacuum_job=latest_vacuum_job,
    )

@app.route("/api/vacuum/advisor")
def api_vacuum_advisor():
    """Ассистент: рекомендации VACUUM/ANALYZE по статистике таблиц."""
    connection_id = request.args.get("connection_id", type=int)

    if not connection_id:
        return jsonify({"ok": False, "message": "connection_id обязателен"}), 400

    try:
        from modules.vacuum_advisor import advise

        return jsonify(advise(connection_id))
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:2000]}), 500


@app.route("/api/vacuum/start", methods=["POST"])
def api_vacuum_start():
    data = request.get_json(silent=True) or {}

    connection_id = data.get("connection_id")
    selected_tables = data.get("tables") or []
    action = data.get("action") or "VACUUM_ANALYZE"

    if not connection_id:
        return jsonify(
            {
                "ok": False,
                "message": "connection_id is required",
            }
        ), 400

    if not selected_tables:
        return jsonify(
            {
                "ok": False,
                "message": "Не выбраны таблицы",
            }
        ), 400

    allowed_actions = [
        "VACUUM",
        "VACUUM_FULL",
        "ANALYZE",
        "VACUUM_ANALYZE",
        "VACUUM_FREEZE",
    ]

    action = str(action).upper().strip()

    if action not in allowed_actions:
        return jsonify(
            {
                "ok": False,
                "message": "Unknown action: {}".format(action),
            }
        ), 400

    try:
        unique_tables = []
        seen = set()

        for item in selected_tables:
            schema_name = None
            table_name = None
            item_action = None

            if isinstance(item, dict):
                schema_name = item.get("schema") or item.get("schema_name")
                table_name = item.get("table") or item.get("table_name")
                # своя операция у таблицы (ассистент): валидируем так же
                item_action = str(item.get("action") or "").upper().strip()

            elif isinstance(item, str):
                value = item.strip()

                if "." in value:
                    parts = value.split(".", 1)
                    schema_name = parts[0].strip().strip('"')
                    table_name = parts[1].strip().strip('"')

            if not schema_name or not table_name:
                continue

            key = "{}.{}".format(schema_name, table_name)

            if key in seen:
                continue

            seen.add(key)

            entry = {
                "schema": schema_name,
                "table": table_name,
            }

            if item_action and item_action in allowed_actions:
                entry["action"] = item_action

            unique_tables.append(entry)

        if not unique_tables:
            return jsonify(
                {
                    "ok": False,
                    "message": "Нет корректных таблиц для запуска",
                }
            ), 400

        try:
            workers = int(data.get("workers") or 1)
        except Exception:
            workers = 1

        job_id = create_job(
            job_type="vacuum_analyze",
            connection_id=int(connection_id),
            config={
                "source": "web",
                "action": action,
                "workers": max(1, min(workers, 8)),
                "tables": unique_tables,
            },
        )

        threading.Thread(
            target=run_vacuum_analyze_job,
            args=(job_id,),
            daemon=True,
        ).start()

        return jsonify(
            {
                "ok": True,
                "job_id": job_id,
                "total_items": len(unique_tables),
                "action": action,
            }
        )

    except Exception as e:
        return jsonify(
            {
                "ok": False,
                "message": str(e),
            }
        ), 500


def update_job_status(job_id, status, error_message=None):
    with sqlite_cursor(commit=True) as cur:
        if error_message is not None:
            cur.execute(
                """
                UPDATE jobs
                SET status = ?,
                    error_message = ?
                WHERE id = ?
                """,
                (
                    status,
                    str(error_message),
                    job_id,
                ),
            )
        else:
            cur.execute(
                """
                UPDATE jobs
                SET status = ?
                WHERE id = ?
                """,
                (
                    status,
                    job_id,
                ),
            )


@app.route("/gpcopy")
def gpcopy_page():
    return render_template(
        "gpcopy_pipeline.html",
        connections=list_connections(),
    )


@app.route("/gpcopy/classic")
def gpcopy_classic_page():
    # Старый интерфейс — оставлен на переходный период.
    return render_template(
        "gpcopy.html",
        connections=list_connections(),
    )


@app.route("/api/gpcopy/start", methods=["POST"])
def api_gpcopy_start():
    data = request.get_json(silent=True) or {}

    source_connection_id = data.get("source_connection_id")
    dest_connection_id = data.get("dest_connection_id")
    selected_tables = data.get("tables") or []

    gpcopy_path = data.get("gpcopy_path") or "/usr/local/gpdb/greenplum-db/bin/gpcopy"
    jobs = data.get("jobs") or 4
    on_segment_threshold = data.get("on_segment_threshold")
    extra_args = data.get("extra_args") or ""

    target_schema = data.get("target_schema") or ""
    target_table_mode = data.get("target_table_mode") or "same"

    truncate = bool(data.get("truncate"))
    drop = bool(data.get("drop"))
    append = bool(data.get("append"))
    skip_existing = bool(data.get("skip_existing"))
    analyze = bool(data.get("analyze"))
    dry_run = bool(data.get("dry_run"))
    validate_count = bool(data.get("validate_count"))

    if not source_connection_id:
        return jsonify({
            "ok": False,
            "message": "source_connection_id is required",
        }), 400

    if not dest_connection_id:
        return jsonify({
            "ok": False,
            "message": "dest_connection_id is required",
        }), 400

    if not selected_tables:
        return jsonify({
            "ok": False,
            "message": "Не выбраны таблицы",
        }), 400

    try:
        source_connection_id = int(source_connection_id)
        dest_connection_id = int(dest_connection_id)
    except Exception:
        return jsonify({
            "ok": False,
            "message": "connection_id должен быть числом",
        }), 400

    # ---------------------------------------------------------
    # ВАЖНО:
    # Убираем дубли выбранных таблиц.
    # Иногда frontend отправляет одну и ту же таблицу 2 раза:
    # например при выборе schema checkbox + table checkbox.
    # ---------------------------------------------------------
    unique_tables = []
    seen = set()

    for item in selected_tables:
        # На случай если frontend отправил строку "schema.table"
        if isinstance(item, str):
            parts = item.split(".", 1)
            if len(parts) != 2:
                continue

            schema_name = parts[0].strip()
            table_name = parts[1].strip()
        else:
            schema_name = item.get("schema") or item.get("schema_name")
            table_name = item.get("table") or item.get("table_name")

        if not schema_name or not table_name:
            continue

        key = (schema_name, table_name)

        if key in seen:
            continue

        seen.add(key)

        unique_tables.append({
            "schema": schema_name,
            "table": table_name,
        })

    if not unique_tables:
        return jsonify({
            "ok": False,
            "message": "После очистки дублей не осталось таблиц для gpcopy",
        }), 400

    try:
        config = {
            "source_connection_id": source_connection_id,
            "dest_connection_id": dest_connection_id,
            "selected_tables": selected_tables,
            "expanded_tables": unique_tables,

            "gpcopy_path": gpcopy_path,
            "jobs": int(jobs),
            "on_segment_threshold": on_segment_threshold,
            "extra_args": extra_args,

            "target_schema": target_schema,
            "target_table_mode": target_table_mode,

            "truncate": truncate,
            "drop": drop,
            "append": append,
            "skip_existing": skip_existing,
            "analyze": analyze,
            "dry_run": dry_run,
            "validate_count": validate_count,
        }

        # выбор транспорта по типам СУБД: gpcopy для GP→GP,
        # copy_pipe (COPY-стрим) для PG↔PG / PG↔GP
        from modules.connections import get_connection_by_id as _conn_cfg

        src_cfg = _conn_cfg(source_connection_id) or {}
        dst_cfg = _conn_cfg(dest_connection_id) or {}
        transport = pick_transport(
            src_cfg.get("db_type"), dst_cfg.get("db_type")
        )

        if transport == "copy_pipe":
            job_type = "copy_pipe"
            runner = run_copy_pipe_job
            action = "COPY"
        else:
            job_type = "gpcopy"
            runner = run_gpcopy_job
            action = "GPCOPY"

        job_id = create_job(
            job_type=job_type,
            connection_id=source_connection_id,
            config=config,
        )

        create_job_items(
            job_id=job_id,
            items=[
                {
                    "schema_name": item["schema"],
                    "table_name": item["table"],
                    "action": action,
                }
                for item in unique_tables
            ],
        )

        threading.Thread(
            target=runner,
            args=(job_id,),
            daemon=True,
        ).start()

        return jsonify({
            "ok": True,
            "job_id": job_id,
            "transport": transport,
            "total_items": len(unique_tables),
            "message": "%s job started" % job_type,
        })

    except ValueError as e:
        return jsonify({
            "ok": False,
            "message": str(e),
        }), 400

    except Exception as e:
        return jsonify({
            "ok": False,
            "message": str(e),
        }), 500

    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500

@app.route("/api/dashboard/session-limits")
def api_dashboard_session_limits():
    connection_id = request.args.get("connection_id", type=int)

    if not connection_id:
        return jsonify({
            "ok": False,
            "message": "connection_id is required",
        }), 400

    try:
        data = get_session_limits_stats(connection_id)

        return jsonify({
            "ok": True,
            "summary": data["summary"],
            "rows": data["rows"],
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "message": str(e),
        }), 500

    except Exception as e:
        return jsonify({
            "ok": False,
            "message": str(e),
        }), 500

@app.route("/api/gpcopy/date-columns")
def api_gpcopy_date_columns():
    connection_id = request.args.get("connection_id", type=int)
    schema_name = request.args.get("schema")
    table_name = request.args.get("table")

    if not connection_id:
        return jsonify({
            "ok": False,
            "message": "connection_id обязателен",
        }), 400

    if not schema_name or not table_name:
        return jsonify({
            "ok": False,
            "message": "schema and table are required",
        }), 400

    try:
        columns = get_gpcopy_date_columns(
            connection_id=connection_id,
            schema_name=schema_name,
            table_name=table_name,
        )

        return jsonify({
            "ok": True,
            "columns": columns,
        })

    except Exception as e:
        import traceback

        return jsonify({
            "ok": False,
            "message": str(e),
            "traceback": traceback.format_exc(),
        }), 500

@app.route("/api/gpcopy/preview-date-json", methods=["POST"])
def api_gpcopy_preview_date_json():
    data = request.get_json(silent=True) or {}

    try:
        include_json = build_gpcopy_date_include_json_preview(data)

        return jsonify({
            "ok": True,
            "include_json": include_json,
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "message": str(e),
        }), 400

@app.route("/api/gpcopy/start-date", methods=["POST"])
def api_gpcopy_start_date():
    data = request.get_json(silent=True) or {}

    source_connection_id = (
        data.get("source_connection_id")
        or data.get("connection_id")
    )

    dest_connection_id = (
        data.get("dest_connection_id")
        or data.get("destination_connection_id")
        or data.get("target_connection_id")
    )

    table_configs = (
        data.get("table_configs")
        or data.get("tables")
        or []
    )

    date_from = data.get("date_from")
    date_to = data.get("date_to")

    if not source_connection_id:
        return jsonify({
            "ok": False,
            "message": "source_connection_id is required",
        }), 400

    if not dest_connection_id:
        return jsonify({
            "ok": False,
            "message": "dest_connection_id is required",
        }), 400

    if not table_configs:
        return jsonify({
            "ok": False,
            "message": "table_configs is empty",
        }), 400

    normalized_tables = []

    for item in table_configs:
        source = item.get("source")
        dest = item.get("dest") or item.get("target")
        sql = item.get("sql")

        schema_name = (
            item.get("schema")
            or item.get("schema_name")
            or item.get("source_schema")
        )

        table_name = (
            item.get("table")
            or item.get("table_name")
            or item.get("source_table")
        )

        date_column = item.get("date_column")

        if not source:
            if schema_name and table_name:
                source = '{}.{}'.format(schema_name, table_name)

        if not dest:
            dest = source

        if not source:
            return jsonify({
                "ok": False,
                "message": "source is required for one table",
                "item": item,
            }), 400

        if not dest:
            return jsonify({
                "ok": False,
                "message": "dest is required for {}".format(source),
                "item": item,
            }), 400

        if not sql:
            if not date_column:
                return jsonify({
                    "ok": False,
                    "message": "date_column is required for {}".format(source),
                    "item": item,
                }), 400

            if not date_from or not date_to:
                return jsonify({
                    "ok": False,
                    "message": "date_from/date_to are required",
                    "item": item,
                }), 400

            sql = "SELECT * FROM {} WHERE {} >= '{}' AND {} < '{}'".format(
                source,
                date_column,
                date_from,
                date_column,
                date_to,
            )

        normalized_tables.append({
            "source": source,
            "dest": dest,
            "sql": sql,
            "schema": schema_name,
            "table": table_name,
            "date_column": date_column,
        })

    try:
        # gpcopy требует ровно один из флагов skip-existing/truncate/drop/append.
        # Для среза по датам дефолт — append (догрузка периода, остальные
        # данные назначения не трогаем).
        flag_chosen = any(
            bool(data.get(k))
            for k in ("append", "truncate", "drop", "skip_existing")
        )

        job_id = create_job(
            job_type="gpcopy",
            connection_id=int(source_connection_id),
            config={
                # run_gpcopy_job включает JSON-срезы только на mode="date_filter"
                "mode": "date_filter",
                "source_connection_id": int(source_connection_id),
                "dest_connection_id": int(dest_connection_id),
                "destination_connection_id": int(dest_connection_id),

                "table_configs": normalized_tables,

                "date_from": date_from,
                "date_to": date_to,

                "jobs": data.get("jobs") or 4,
                "on_segment_threshold": data.get("on_segment_threshold", -1),

                "append": bool(data.get("append")) or not flag_chosen,
                "truncate": bool(data.get("truncate")),
                "drop": bool(data.get("drop")),
                "skip_existing": bool(data.get("skip_existing")),
                "no_ownership": bool(data.get("no_ownership")),
                "analyze": bool(data.get("analyze")),
                "dry_run": bool(data.get("dry_run")),
            },
        )

        create_job_items(
            job_id=job_id,
            items=[
                {
                    "schema_name": item.get("schema") or item["source"].split(".")[0],
                    "table_name": item.get("table") or item["source"].split(".")[-1],
                    "action": "GPCOPY DATE",
                }
                for item in normalized_tables
            ],
        )

        import threading

        threading.Thread(
            target=run_gpcopy_job,
            args=(job_id,),
            daemon=True,
        ).start()

        return jsonify({
            "ok": True,
            "job_id": job_id,
            "total_items": len(normalized_tables),
            "message": "GPCOPY by date started",
        })

    except Exception as e:
        import traceback

        return jsonify({
            "ok": False,
            "message": str(e),
            "traceback": traceback.format_exc(),
        }), 500

def get_app_sqlite_path():
    # Единый источник пути к БД (config.SQLITE_DB_PATH), подменяемый в тестах.
    return config.SQLITE_DB_PATH


def sqlite_rows(query, params=None):
    conn = sqlite3.connect(get_app_sqlite_path())
    conn.row_factory = sqlite3.Row

    try:
        cur = conn.cursor()
        cur.execute(query, params or [])
        rows = cur.fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()

@app.route("/api/jobs/<int:job_id>/skew-export.xlsx")
def api_export_skew_job_excel(job_id):
    try:
        job_rows = sqlite_rows(
            """
            SELECT *
            FROM jobs
            WHERE id = ?
            """,
            [job_id]
        )

        if not job_rows:
            return jsonify({
                "ok": False,
                "message": "Job not found",
            }), 404

        job = job_rows[0]

        items = sqlite_rows(
            """
            SELECT *
            FROM job_items
            WHERE job_id = ?
            ORDER BY id
            """,
            [job_id]
        )

        results = sqlite_rows(
            """
            SELECT *
            FROM skew_results
            WHERE job_id = ?
            ORDER BY schema_name, table_name
            """,
            [job_id]
        )

        result_map = {}

        for row in results:
            key = "{}.{}".format(
                row.get("schema_name"),
                row.get("table_name")
            )
            result_map[key] = row

        wb = Workbook()
        ws = wb.active
        ws.title = "Skew Analysis"

        headers = [
            "Job ID",
            "Schema",
            "Table",
            "Item Status",
            "Skew Status",
            "Skew Ratio",
            "Total Rows",
            "Segment Count",
            "Empty Segments",
            "Max Rows",
            "Min Rows",
            "Duration Seconds",
            "Error",
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for item in items:
            schema_name = (
                item.get("schema_name")
                or item.get("schema")
                or ""
            )

            table_name = (
                item.get("table_name")
                or item.get("table")
                or ""
            )

            key = "{}.{}".format(schema_name, table_name)
            skew = result_map.get(key, {})

            total_rows = skew.get("total_rows")
            empty_segments = skew.get("empty_segments")
            segment_count = (
                skew.get("segment_count")
                or skew.get("segments_count")
                or skew.get("total_segments")
                or ""
            )

            ws.append([
                job_id,
                schema_name,
                table_name,
                item.get("status") or "",
                skew.get("status") or "",
                skew.get("skew_ratio") or "",
                total_rows if total_rows is not None else "",
                segment_count,
                empty_segments if empty_segments is not None else "",
                skew.get("max_rows") or "",
                skew.get("min_rows") or "",
                item.get("duration_seconds") or "",
                item.get("error_message") or "",
            ])

        # Цвета по статусам
        status_colors = {
            "done": "C6EFCE",
            "failed": "FFC7CE",
            "running": "BDD7EE",
            "queued": "D9EAD3",
            "skipped": "FFF2CC",
            "OK": "C6EFCE",
            "WARNING": "FFF2CC",
            "CRITICAL": "FFC7CE",
            "EMPTY": "D9EAD3",
            "FAILED": "FFC7CE",
        }

        for row in ws.iter_rows(min_row=2):
            item_status_cell = row[3]
            skew_status_cell = row[4]

            for cell in (item_status_cell, skew_status_cell):
                color = status_colors.get(str(cell.value))
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)

        # Автоширина
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 60)

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "skew_analysis_job_{}.xlsx".format(job_id)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        import traceback

        return jsonify({
            "ok": False,
            "message": str(e),
            "traceback": traceback.format_exc(),
        }), 500

@app.route("/api/gpcopy/sync/preview", methods=["POST"])
def api_gpcopy_sync_preview():
    data = request.get_json(silent=True) or {}

    try:
        result = preview_gpcopy_sync(data)
        return jsonify(result)

    except Exception as e:
        import traceback
        return jsonify({
            "ok": False,
            "message": str(e),
            "traceback": traceback.format_exc(),
        }), 500


@app.route("/api/gpcopy/sync/apply", methods=["POST"])
def api_gpcopy_sync_apply():
    data = request.get_json(silent=True) or {}

    source_connection_id = data.get("source_connection_id")
    dest_connection_id = data.get("dest_connection_id")
    table_configs = data.get("table_configs") or []

    if not source_connection_id:
        return jsonify({"ok": False, "message": "source_connection_id is required"}), 400

    if not dest_connection_id:
        return jsonify({"ok": False, "message": "dest_connection_id is required"}), 400

    if not table_configs:
        return jsonify({"ok": False, "message": "table_configs is empty"}), 400

    try:
        job_id = create_job(
            job_type="gpcopy_sync",
            connection_id=int(source_connection_id),
            config={
                "mode": "sync_diff",
                "source_connection_id": int(source_connection_id),
                "dest_connection_id": int(dest_connection_id),
                "table_configs": table_configs,
                "gpcopy_path": data.get("gpcopy_path"),
                "jobs": data.get("jobs") or 4,
            },
        )

        create_job_items(
            job_id=job_id,
            items=[
                {
                    "schema_name": cfg.get("schema"),
                    "table_name": cfg.get("table"),
                    "action": "GPCOPY SYNC",
                }
                for cfg in table_configs
            ],
        )

        import threading

        threading.Thread(
            target=run_gpcopy_sync_job,
            args=(job_id,),
            daemon=True,
        ).start()

        return jsonify({
            "ok": True,
            "job_id": job_id,
            "message": "GPCOPY sync started",
        })

    except Exception as e:
        import traceback
        return jsonify({
            "ok": False,
            "message": str(e),
            "traceback": traceback.format_exc(),
        }), 500

@app.route("/health")
def health_page():
    return render_template(
        "health.html",
        connections=list_connections(),
    )


@app.route("/backups")
def backups_page():
    return render_template(
        "backups.html",
        connections=list_connections(),
    )


@app.route("/api/backup/list")
def api_backup_list():
    return jsonify({"ok": True, "backups": list_backup_records()})


@app.route("/api/backup/start", methods=["POST"])
def api_backup_start():
    data = request.get_json(silent=True) or {}
    connection_id = data.get("connection_id")

    if not connection_id:
        return jsonify({"ok": False, "message": "connection_id обязателен"}), 400

    config = {
        "connection_id": int(connection_id),
        "backup_type": data.get("backup_type") or "full",
        "backup_dir": (data.get("backup_dir") or "").strip(),
        "include_schemas": data.get("include_schemas") or "",
        "include_tables": data.get("include_tables") or "",
        "jobs": data.get("jobs") or 1,
        "compression_level": data.get("compression_level"),
        "gpbackup_path": (data.get("gpbackup_path") or "").strip(),
        "extra_args": data.get("extra_args") or "",
    }

    try:
        # валидация конфига до создания задачи (понятная ошибка сразу)
        from modules.connections import get_connection_by_id as _cfg
        from modules.gpbackup import build_gpbackup_command

        conn_cfg = _cfg(int(connection_id))

        if not conn_cfg:
            return jsonify({"ok": False, "message": "Подключение не найдено"}), 400

        probe = dict(config)
        probe.setdefault(
            "dbname",
            conn_cfg.get("database_name") or conn_cfg.get("database"),
        )
        build_gpbackup_command(probe)
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400

    job_id = create_job(
        job_type="gpbackup",
        connection_id=int(connection_id),
        config=config,
    )

    create_job_items(
        job_id=job_id,
        items=[{
            "schema_name": conn_cfg.get("database_name") or "",
            "table_name": "gpbackup ({})".format(config["backup_type"]),
            "action": "GPBACKUP",
        }],
    )

    threading.Thread(target=run_gpbackup_job, args=(job_id,), daemon=True).start()

    return jsonify({"ok": True, "job_id": job_id, "message": "Бэкап запущен"})


@app.route("/api/backup/restore", methods=["POST"])
def api_backup_restore():
    data = request.get_json(silent=True) or {}
    connection_id = data.get("connection_id")
    backup_timestamp = data.get("backup_timestamp")

    if not connection_id or not backup_timestamp:
        return jsonify({
            "ok": False,
            "message": "connection_id и backup_timestamp обязательны",
        }), 400

    config = {
        "connection_id": int(connection_id),
        "backup_timestamp": str(backup_timestamp),
        "backup_dir": (data.get("backup_dir") or "").strip(),
        "redirect_db": (data.get("redirect_db") or "").strip(),
        "create_db": bool(data.get("create_db")),
        "data_only": bool(data.get("data_only")),
        "include_tables": data.get("include_tables") or "",
        "jobs": data.get("jobs") or 1,
        "gprestore_path": (data.get("gprestore_path") or "").strip(),
    }

    try:
        from modules.gpbackup import build_gprestore_command

        build_gprestore_command(config)
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400

    job_id = create_job(
        job_type="gprestore",
        connection_id=int(connection_id),
        config=config,
    )

    create_job_items(
        job_id=job_id,
        items=[{
            "schema_name": config["redirect_db"] or "",
            "table_name": "gprestore {}".format(backup_timestamp),
            "action": "GPRESTORE",
        }],
    )

    threading.Thread(target=run_gprestore_job, args=(job_id,), daemon=True).start()

    return jsonify({"ok": True, "job_id": job_id, "message": "Восстановление запущено"})


@app.route("/api/backup/sync-disk", methods=["POST"])
def api_backup_sync_disk():
    """Реестр <- диск: gpbackup_manager list-backups с координатора."""
    data = request.get_json(silent=True) or {}
    connection_id = data.get("connection_id")

    if not connection_id:
        return jsonify({"ok": False, "message": "connection_id обязателен"}), 400

    try:
        from modules.gpbackup import sync_disk_backups

        result = sync_disk_backups(
            int(connection_id),
            manager_path=(data.get("gpbackup_manager_path") or "").strip() or None,
        )
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:2000]}), 500


@app.route("/api/backup/report", methods=["POST"])
def api_backup_report():
    data = request.get_json(silent=True) or {}
    connection_id = data.get("connection_id")
    backup_timestamp = data.get("backup_timestamp")

    if not connection_id or not backup_timestamp:
        return jsonify({
            "ok": False,
            "message": "connection_id и backup_timestamp обязательны",
        }), 400

    try:
        from modules.gpbackup import get_backup_report

        report = get_backup_report(int(connection_id), str(backup_timestamp))
        return jsonify({"ok": True, "report": report[-20000:]})
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:2000]}), 500


@app.route("/api/backup/delete", methods=["POST"])
def api_backup_delete():
    data = request.get_json(silent=True) or {}
    connection_id = data.get("connection_id")
    backup_timestamp = data.get("backup_timestamp")

    if not connection_id or not backup_timestamp:
        return jsonify({
            "ok": False,
            "message": "connection_id и backup_timestamp обязательны",
        }), 400

    try:
        from modules.gpbackup import delete_disk_backup

        delete_disk_backup(int(connection_id), str(backup_timestamp))
        return jsonify({"ok": True, "message": "Копия удалена с диска"})
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)[:2000]}), 500


@app.route("/api/health/overview")
def api_health_overview():
    connection_id = request.args.get("connection_id", type=int)
    force = request.args.get("force") == "1"

    if not connection_id:
        return jsonify({"ok": False, "message": "connection_id обязателен"}), 400

    try:
        from modules.db_health import collect_health
        return jsonify(collect_health(connection_id, force=force))
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


# ============================================================
# Scheduler (spec §8)
# ============================================================

@app.route("/schedules")
def schedules_page():
    return render_template(
        "schedules.html",
        connections=list_connections(),
    )


@app.route("/api/schedules", methods=["GET", "POST"])
def api_schedules():
    if request.method == "GET":
        return jsonify({"ok": True, "schedules": scheduler_store.list_schedules()})

    data = request.get_json(silent=True) or {}

    try:
        schedule_id = scheduler_store.create_schedule({
            "name": data.get("name"),
            "job_type": data.get("job_type"),
            "cron_expr": data.get("cron_expr"),
            "config_json": _json.dumps(data.get("config") or {}, ensure_ascii=False),
            "overlap_policy": data.get("overlap_policy"),
            "max_retries": data.get("max_retries"),
            "retry_delay_seconds": data.get("retry_delay_seconds"),
            "notify_on": data.get("notify_on"),
            "notify_channel_ids": _json.dumps(data.get("notify_channel_ids") or []),
            "enabled": data.get("enabled", 1),
        })
        return jsonify({"ok": True, "id": schedule_id})
    except (ValueError, KeyError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400


@app.route("/api/schedules/<int:schedule_id>", methods=["PUT", "DELETE"])
def api_schedule_item(schedule_id):
    if request.method == "DELETE":
        scheduler_store.delete_schedule(schedule_id)
        return jsonify({"ok": True})

    data = request.get_json(silent=True) or {}
    update = {}

    for key in (
        "name", "job_type", "cron_expr", "overlap_policy",
        "max_retries", "retry_delay_seconds", "notify_on", "enabled",
    ):
        if key in data:
            update[key] = data[key]

    if "config" in data:
        update["config_json"] = _json.dumps(data["config"], ensure_ascii=False)

    if "notify_channel_ids" in data:
        update["notify_channel_ids"] = _json.dumps(data["notify_channel_ids"])

    try:
        scheduler_store.update_schedule(schedule_id, update)
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 400


@app.route("/api/schedules/<int:schedule_id>/toggle", methods=["POST"])
def api_schedule_toggle(schedule_id):
    schedule = scheduler_store.get_schedule(schedule_id)

    if not schedule:
        return jsonify({"ok": False, "message": "Schedule not found"}), 404

    new_enabled = 0 if schedule["enabled"] else 1
    scheduler_store.set_enabled(schedule_id, new_enabled)
    return jsonify({"ok": True, "enabled": new_enabled})


@app.route("/api/schedules/<int:schedule_id>/run-now", methods=["POST"])
def api_schedule_run_now(schedule_id):
    try:
        result = gpm_scheduler.run_now(schedule_id)
        return jsonify({"ok": True, **result})
    except ValueError as e:
        return jsonify({"ok": False, "message": str(e)}), 404


@app.route("/api/schedules/<int:schedule_id>/runs")
def api_schedule_runs(schedule_id):
    return jsonify({"ok": True, "runs": scheduler_store.list_runs(schedule_id)})


@app.route("/api/schedules/preview", methods=["POST"])
def api_schedules_preview():
    data = request.get_json(silent=True) or {}
    out = {"ok": True}

    cron_expr = data.get("cron_expr")

    if cron_expr:
        if not scheduler_store.validate_cron(cron_expr):
            return jsonify({"ok": False, "message": "Invalid cron expression"}), 400

        from croniter import croniter as _croniter

        it = _croniter(cron_expr, _datetime.now())
        out["next_runs"] = [
            it.get_next(_datetime).strftime("%Y-%m-%d %H:%M:%S") for _ in range(5)
        ]

    date_window = data.get("date_window")

    if date_window:
        try:
            date_from, date_to = resolve_date_window(date_window, _datetime.now())
            out["date_from"] = date_from
            out["date_to"] = date_to
        except ValueError as e:
            return jsonify({"ok": False, "message": str(e)}), 400

    return jsonify(out)


@app.route("/api/notification-channels", methods=["GET", "POST"])
def api_notification_channels():
    if request.method == "GET":
        return jsonify({"ok": True, "channels": scheduler_store.list_channels()})

    data = request.get_json(silent=True) or {}

    try:
        channel_id = scheduler_store.create_channel({
            "name": data.get("name"),
            "type": data.get("type") or "webhook",
            "config_json": _json.dumps(data.get("config") or {}, ensure_ascii=False),
            "enabled": data.get("enabled", 1),
        })
        return jsonify({"ok": True, "id": channel_id})
    except KeyError as e:
        return jsonify({"ok": False, "message": str(e)}), 400


@app.route("/api/notification-channels/<int:channel_id>", methods=["PUT", "DELETE"])
def api_notification_channel_item(channel_id):
    if request.method == "DELETE":
        scheduler_store.delete_channel(channel_id)
        return jsonify({"ok": True})

    data = request.get_json(silent=True) or {}
    update = {}

    for key in ("name", "type", "enabled"):
        if key in data:
            update[key] = data[key]

    if "config" in data:
        update["config_json"] = _json.dumps(data["config"], ensure_ascii=False)

    scheduler_store.update_channel(channel_id, update)
    return jsonify({"ok": True})


# ============================================================
# Table catalog: массовый выбор/настройка (10k+ таблиц)
# ============================================================

import modules.table_catalog as table_catalog


@app.route("/api/catalog")
def api_catalog():
    connection_id = request.args.get("connection_id", type=int)
    force = request.args.get("force", 0, type=int)

    if not connection_id:
        return jsonify({"ok": False, "message": "connection_id обязателен"}), 400

    try:
        tables, cached_at = table_catalog.get_catalog(connection_id, force=bool(force))
        return jsonify({
            "ok": True,
            "total": len(tables),
            "cached_at": cached_at,
            "schemas": table_catalog.catalog_summary(tables),
        })
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/catalog/search")
def api_catalog_search():
    connection_id = request.args.get("connection_id", type=int)
    query = request.args.get("q", "")

    try:
        tables, _ts = table_catalog.get_catalog(connection_id)
        found = table_catalog.search_tables(tables, query)
        return jsonify({
            "ok": True,
            "tables": [{"schema": s, "table": t} for s, t in found],
        })
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/catalog/schema-tables")
def api_catalog_schema_tables():
    connection_id = request.args.get("connection_id", type=int)
    schema = (request.args.get("schema") or "").strip()

    if not connection_id or not schema:
        return jsonify({"ok": False, "message": "connection_id и schema обязательны"}), 400

    try:
        return jsonify({
            "ok": True,
            "tables": table_catalog.schema_tables_with_roles(connection_id, schema),
        })
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/catalog/expand-mask", methods=["POST"])
def api_catalog_expand_mask():
    data = request.get_json(silent=True) or {}

    try:
        tables, _ts = table_catalog.get_catalog(int(data["connection_id"]))
        matched = table_catalog.match_mask(tables, data.get("mask"))
        return jsonify({
            "ok": True,
            "count": len(matched),
            "tables": [{"schema": s, "table": t} for s, t in matched],
        })
    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/catalog/resolve-list", methods=["POST"])
def api_catalog_resolve_list():
    data = request.get_json(silent=True) or {}

    try:
        tables, _ts = table_catalog.get_catalog(int(data["connection_id"]))
        valid, invalid = table_catalog.parse_table_list(data.get("text"), tables)
        return jsonify({
            "ok": True,
            "valid": [{"schema": s, "table": t} for s, t in valid],
            "invalid": invalid,
        })
    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/catalog/resolve-columns", methods=["POST"])
def api_catalog_resolve_columns():
    data = request.get_json(silent=True) or {}

    try:
        tables = [
            (t["schema"], t["table"])
            for t in (data.get("tables") or [])
        ]
        priority = [
            c.strip() for c in (data.get("priority") or []) if str(c).strip()
        ]

        if not tables or not priority:
            return jsonify({
                "ok": False, "message": "tables и priority обязательны",
            }), 400

        columns_by_table = table_catalog.fetch_columns_for_candidates(
            int(data["connection_id"]), tables, priority,
        )

        if data.get("fallback_any_date"):
            # Фолбэк: у таблиц без колонок из приоритета берём первую
            # date/timestamp колонку (не у всех таблиц колонки зовутся одинаково).
            _base_resolved, base_missing = table_catalog.pick_columns(
                columns_by_table, priority,
            )
            date_cols = table_catalog.fetch_date_columns_bulk(
                int(data["connection_id"]), base_missing,
            )
            resolved, missing = table_catalog.pick_columns_with_fallback(
                columns_by_table, priority, date_cols,
            )

            return jsonify({
                "ok": True,
                "resolved": [
                    {"schema": s, "table": t,
                     "column": info["column"], "via": info["via"]}
                    for (s, t), info in sorted(resolved.items())
                ],
                "missing": [{"schema": s, "table": t} for s, t in missing],
            })

        resolved, missing = table_catalog.pick_columns(columns_by_table, priority)

        return jsonify({
            "ok": True,
            "resolved": [
                {"schema": s, "table": t, "column": c, "via": "priority"}
                for (s, t), c in sorted(resolved.items())
            ],
            "missing": [{"schema": s, "table": t} for s, t in missing],
        })
    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/catalog/primary-keys", methods=["POST"])
def api_catalog_primary_keys():
    data = request.get_json(silent=True) or {}

    try:
        tables = [(t["schema"], t["table"]) for t in (data.get("tables") or [])]
        keys = table_catalog.fetch_primary_keys(int(data["connection_id"]), tables)

        return jsonify({
            "ok": True,
            "keys": [
                {"schema": s, "table": t, "columns": cols}
                for (s, t), cols in sorted(keys.items())
            ],
            "missing": [
                {"schema": s, "table": t}
                for s, t in tables if (s, t) not in keys
            ],
        })
    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/catalog/resolve-keys", methods=["POST"])
def api_catalog_resolve_keys():
    """Ключи для sync: PK -> уникальный индекс -> unresolved (на вычисление)."""
    data = request.get_json(silent=True) or {}

    try:
        tables = [(t["schema"], t["table"]) for t in (data.get("tables") or [])]

        if not tables:
            return jsonify({"ok": False, "message": "tables is empty"}), 400

        pk_map, unique_map = table_catalog.fetch_unique_indexes(
            int(data["connection_id"]), tables,
        )
        resolved, unresolved = table_catalog.resolve_keys_hierarchy(
            tables, pk_map, unique_map,
        )

        return jsonify({
            "ok": True,
            "resolved": [
                {"schema": s, "table": t,
                 "columns": info["columns"], "source": info["source"]}
                for (s, t), info in sorted(resolved.items())
            ],
            "unresolved": [{"schema": s, "table": t} for s, t in unresolved],
        })
    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/catalog/compute-unique", methods=["POST"])
def api_catalog_compute_unique():
    """Вычисление уникальной колонки по данным (нет ни PK, ни индексов)."""
    data = request.get_json(silent=True) or {}

    try:
        result = table_catalog.probe_unique_column(
            int(data["connection_id"]),
            data["schema"],
            data["table"],
            limit_candidates=int(data.get("limit_candidates") or 5),
        )
        return jsonify({"ok": True, **result})
    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/table-sets", methods=["GET", "POST"])
def api_table_sets():
    if request.method == "GET":
        connection_id = request.args.get("connection_id", type=int)
        return jsonify({
            "ok": True,
            "sets": table_catalog.list_table_sets(connection_id),
        })

    data = request.get_json(silent=True) or {}

    if not data.get("tables"):
        return jsonify({"ok": False, "message": "tables is empty"}), 400

    set_id = table_catalog.create_table_set(data)
    return jsonify({"ok": True, "id": set_id})


@app.route("/api/table-sets/<int:set_id>", methods=["GET", "DELETE"])
def api_table_set_item(set_id):
    if request.method == "DELETE":
        table_catalog.delete_table_set(set_id)
        return jsonify({"ok": True})

    ts = table_catalog.get_table_set(set_id)

    if not ts:
        return jsonify({"ok": False, "message": "Set not found"}), 404

    return jsonify({"ok": True, "set": ts})


# ============================================================
# gpcopy v2: increment + partition-diff
# ============================================================

@app.route("/api/gpcopy/increment/preview", methods=["POST"])
def api_gpcopy_increment_preview():
    data = request.get_json(silent=True) or {}

    try:
        from db import get_connection_by_id as _get_conn

        dest_cfg = _get_conn(int(data["dest_connection_id"]))
        source_cfg = _get_conn(int(data["source_connection_id"]))
        tables = data.get("tables") or []

        if not tables:
            return jsonify({"ok": False, "message": "tables is empty"}), 400

        watermarks = {}
        preview = []

        for entry in tables:
            schema = entry.get("schema")
            table = entry.get("table")
            column = entry.get("watermark_column")

            if not (schema and table and column):
                return jsonify({
                    "ok": False,
                    "message": "schema/table/watermark_column обязательны",
                }), 400

            wm = get_dest_watermark(dest_cfg, schema, table, column)
            watermarks[(schema, table)] = wm
            preview.append({
                "schema": schema,
                "table": table,
                "watermark_column": column,
                "watermark": str(wm) if wm is not None else None,
            })

        items = build_increment_items(
            tables, watermarks,
            data.get("source_db") or "src", data.get("dest_db") or "dst",
        )

        for row, item in zip(preview, items):
            row["sql"] = item["sql"]

        return jsonify({"ok": True, "tables": preview})

    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/gpcopy/increment/start", methods=["POST"])
def api_gpcopy_increment_start():
    data = request.get_json(silent=True) or {}

    try:
        tables = data.get("tables") or []

        if not tables:
            return jsonify({"ok": False, "message": "tables is empty"}), 400

        job_id = create_job(
            job_type="gpcopy_increment",
            connection_id=int(data["source_connection_id"]),
            config=data,
        )

        threading.Thread(
            target=run_gpcopy_increment_job, args=(job_id,), daemon=True
        ).start()

        return jsonify({"ok": True, "job_id": job_id})

    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400


@app.route("/api/gpcopy/partition-diff/preview", methods=["POST"])
def api_gpcopy_partition_diff_preview():
    data = request.get_json(silent=True) or {}

    try:
        from db import get_connection_by_id as _get_conn

        source_cfg = _get_conn(int(data["source_connection_id"]))
        dest_cfg = _get_conn(int(data["dest_connection_id"]))

        schema = data.get("schema")
        table = data.get("table")

        if not (schema and table):
            return jsonify({"ok": False, "message": "schema/table обязательны"}), 400

        diff_rows, _leaves = diff_partitions(source_cfg, dest_cfg, schema, table)

        return jsonify({
            "ok": True,
            "partitions": diff_rows,
            "to_copy": [r for r in diff_rows if r["action"] != "skip"],
        })

    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/gpcopy/partition-diff/preview-bulk", methods=["POST"])
def api_gpcopy_partition_diff_preview_bulk():
    """Быстрый diff по статистике каталога для пачки таблиц (один запрос на сторону)."""
    data = request.get_json(silent=True) or {}

    try:
        from db import get_connection_by_id as _get_conn
        from modules.gpcopy_partition import diff_partitions_stats

        source_cfg = _get_conn(int(data["source_connection_id"]))
        dest_cfg = _get_conn(int(data["dest_connection_id"]))

        tables = [
            (t["schema"], t["table"])
            for t in (data.get("tables") or [])
        ]
        if not tables:
            return jsonify({"ok": False, "message": "tables is empty"}), 400

        diff_by_root, leaves_by_root = diff_partitions_stats(
            source_cfg, dest_cfg, tables,
            exact=bool(data.get("exact")),
        )

        out = []
        total_copy = 0
        for (schema, table) in tables:
            rows = diff_by_root.get((schema, table)) or []
            leaves = leaves_by_root.get((schema, table)) or {}
            missing = sum(1 for r in rows if r["action"] == "copy_missing")
            changed = sum(1 for r in rows if r["action"] == "copy_changed")
            total_copy += missing + changed

            # детализация по ВСЕМ партициям — фильтры (все/разл./совпад.) в UI
            detail = [
                {
                    "partition": r["partition"],
                    "schema": (leaves.get(r["partition"]) or (schema,))[0],
                    "src": r["src_count"],
                    "dest": r["dest_count"],
                    "action": r["action"],
                }
                for r in rows
            ]

            out.append({
                "schema": schema,
                "table": table,
                "partitions": len(rows),
                "missing": missing,
                "changed": changed,
                "to_copy": missing + changed,
                "detail": detail,
            })

        return jsonify({"ok": True, "tables": out, "total_to_copy": total_copy})

    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500


@app.route("/api/gpcopy/partition-diff/start", methods=["POST"])
def api_gpcopy_partition_diff_start():
    data = request.get_json(silent=True) or {}

    try:
        tables = data.get("tables") or []

        if not tables:
            return jsonify({"ok": False, "message": "tables is empty"}), 400

        job_id = create_job(
            job_type="gpcopy_partition_diff",
            connection_id=int(data["source_connection_id"]),
            config=data,
        )

        threading.Thread(
            target=run_gpcopy_partition_diff_job, args=(job_id,), daemon=True
        ).start()

        return jsonify({"ok": True, "job_id": job_id})

    except (KeyError, ValueError) as e:
        return jsonify({"ok": False, "message": str(e)}), 400


@app.route("/api/notification-channels/<int:channel_id>/test", methods=["POST"])
def api_notification_channel_test(channel_id):
    channel = scheduler_store.get_channel(channel_id)

    if not channel:
        return jsonify({"ok": False, "message": "Channel not found"}), 404

    import notifiers as _notifiers

    ok, error = _notifiers.send(channel, {
        "schedule": "test-event",
        "job_type": "test",
        "status": "test",
        "fired_at": _datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "error": None,
        "job_id": None,
    })

    return jsonify({"ok": ok, "message": error})


if __name__ == "__main__":
    init_db()

    # переподхват после рестарта: внешние бинари (gpcopy/gpbackup) живут
    # сами — цепляемся к pid и логу; in-process задачи (vacuum, reorganize,
    # skew, sync, copy_pipe) перезапускаем с места остановки
    from modules.gpbackup import resume_unfinished_backup_jobs
    from modules.gpcopy import resume_unfinished_gpcopy_jobs
    from modules.job_resume import resume_inprocess_jobs

    resumed_jobs = resume_unfinished_gpcopy_jobs()
    resumed_jobs += resume_unfinished_backup_jobs()
    resumed_jobs += resume_inprocess_jobs(exclude_ids=resumed_jobs)

    if resumed_jobs:
        print("Resumed jobs after restart:", resumed_jobs)

    interrupted_jobs = mark_interrupted_jobs_on_startup(exclude_ids=resumed_jobs)

    if interrupted_jobs:
        print("Interrupted jobs after application startup:", interrupted_jobs)

    from scheduler import start_scheduler
    start_scheduler()

    app.run(
        host=APP_HOST,
        port=APP_PORT,
        debug=APP_DEBUG,
        use_reloader=False,
    )
